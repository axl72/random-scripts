import pandas as pd 
from openpyxl import load_workbook
from openpyxl.styles import Font
import argparse

DB_PATH = "C:/Users/dberrospi/Desktop/INTEK/MAESTRO PRODUCTOS INTEK.xlsx"
TARGET = "C:/Users/dberrospi/Downloads/ESPECIALES CAJA GRANDE ESTILOS 25.05.2026 CARGA JULIO DDN AGOSTO.xlsx"
class Parser:
    def __init__(self):
        self.parser = argparse.ArgumentParser(description="Inserta links de imágenes y videos en un archivo Excel comparando códigos SKU.")

        self.command_subparsers = self.parser.add_subparsers(dest="commands", help="Comandos disponibles")
        self.set_parser = self.command_subparsers.add_parser("set", help="Establecer coasas")
        self.resource_subparsers =  self.set_parser.add_subparsers(dest="resources", required=True, help="Establecer la base de datos de links desde un archivo Excel.")
        self.database_parser = self.resource_subparsers.add_parser("database", help="Establecer la base de datos de links desde un archivo Excel.")

        self.database_parser.add_argument("database_path", default=None, help="Archivo Excel que contiene la base de datos de links.")
        self.database_parser.add_argument("--id_col_name", default="SKU", help="Nombre de la columna que contiene el identificador (por defecto: SKU).")
        self.database_parser.add_argument("--image_col_name", default="IMAGENES LINK", help="Nombre de la columna que contiene los links de imágenes (por defecto: IMAGENES LINK).")
        self.database_parser.add_argument("--video_col_name", default="VIDEO LINK", help="Nombre de la columna que contiene los links de videos (por defecto: VIDEO LINK).")

        self.parser.add_argument("--input-file", default="", help="Ruta del archivo Excel que será procesado.")
        self.parser.add_argument("--output-file", default="output.xlsx", help="Ruta del archivo Excel de salida.")
        self.parser.add_argument("--also-video", action="store_true", help="Incluir esta opción para insertar también los links de video (por defecto: solo imágenes).")
        self.parser.add_argument("--col-codigo", default="C", help="Columna donde se encuentra el código SKU (por defecto: C).")
        self.parser.add_argument("--col-image-link", default="B", help="Columna donde se insertará el link de la imagen (por defecto: B).")
        self.parser.add_argument("--col-video-link", default="A", help="Columna donde se insertará el link del video (por defecto: A).")
        self.parser.add_argument("--all-sheets", action="store_true", default=False, help="Procesar todas las hojas del archivo de entrada (por defecto: solo la hoja especificada).")
        self.parser.add_argument("--sheet-name", help="Nombre de la hoja")

    def parse_args(self):
        return self.parser.parse_args()
    
class LinkInserter:
    def __init__(self):
        self.database_info = {"db_path": DB_PATH, id: "SKU", "image": "IMAGENES LINK", "video": "VIDEO LINK"}

    def set_databse(self, database_info:dict):
        """Establecer la base de datos implica que debes pasar como argumento la ruta del Excel
        que contiene los links, debes definir el nombre de la columna identificadora,
        el nombre de la columna del valor del link."""
        self.database_info = database_info


    def _get_dict_from_excel(self, ruta_fotos, sheet=0):
        """Esta función lee un archivo Excel y devuelve un diccionario con la estructura {codigo: {image_url: url, video_url: url}}.
        En otras palabras, esta función lee una base de datos registrada en un Excel y la convierte en un diccionario que luego se usará
        para insertar los links en el archivo de destino."""
        df = pd.read_excel(ruta_fotos, sheet_name=sheet, dtype=str)

        # Normalizar columnas
        df.columns = df.columns.str.strip().str.lower()

        # Limpiar valores clave
        df["sku"] = df["sku"].str.strip()

        # Crear estructura
        resultado = (
            df.set_index("sku")[["imagenes link", "video link"]]
            .rename(columns={
                "imagenes link": "image_url",
                "video link": "video_url"
            })
            .to_dict(orient="index")
        )

        return resultado

    def insertar_links_excel(self,
        archivo_entrada:str,
        archivo_salida:str,
        sheet_name:str=None,
        also_video=False,
        col_codigo:str="C", # Columna donde se encuentra el código SKU
        col_image_link:str="B", # Columna donde se insertará el link de la imagen
        col_video_link:str="A", # Columna donde se insertará el link del video
        dic_links:dict[str, dict[str, str]]=None,
        **kwargs
    ):
        """
        Inserta hipervínculos comparando siempre como strings.
        """
        if dic_links is None:
            raise ValueError("Debes pasar un diccionario de links {codigo: {image_url: url}}")

        wb = load_workbook(archivo_entrada)
        print(wb.sheetnames)
        input("Presiona Enter para continuar...")  # Pausa para verificar las hojas disponibles
        ws = wb[wb.sheetnames[0]] if not sheet_name else wb[sheet_name]  # Por defecto, se toma la primera hoja

        for row in range(2, ws.max_row + 1):
            valor_celda = ws[f"{col_codigo}{row}"].value
            
            if valor_celda is None:
                continue

            # PARSEO CRÍTICO: Convertimos el valor de Excel a string y limpiamos espacios
            codigo_excel = str(valor_celda).strip()

            if codigo_excel not in dic_links:
                continue 

            image_url = dic_links[codigo_excel]["image_url"]
            video_url = dic_links[codigo_excel]["video_url"]

            if image_url is not None and str(image_url).strip() and str(image_url).lower() != "nan":
                cell = ws[f"{col_image_link}{row}"]
                cell.value = "Ver Foto"  # Texto que se verá en la celda
                cell.hyperlink = image_url
                cell.font = Font(color="0000FF", underline="single")
            
            if video_url is not None and str(video_url).strip() and str(video_url).lower() != "nan" and also_video:
                cell = ws[f"{col_video_link}{row}"]
                cell.value = "Ver Video"  # Texto que se verá en la celda
                cell.hyperlink = video_url
                cell.font = Font(color="0000FF", underline="single")

        wb.save(archivo_salida)
        print(f"Archivo guardado exitosamente en: {archivo_salida}")

    def main(self):
        parser = Parser()
        args = parser.parse_args()

        if args.commands == "set":

            if args.resources == "database":


                self.set_databse({
                    "db_path": args.database_path,
                    "id": args.id_col_name,
                    "image": args.image_col_name,
                    "video": args.video_col_name
                })
            return
        
        # Flujo cuando se deciden procesar todas las hojas
        if args.all_sheets != 0:
            pass

        links = self._get_dict_from_excel(self.database_info["db_path"])

        print(self.database_info)
        self.insertar_links_excel(
            archivo_entrada=args.input_file,
            archivo_salida=args.output_file,
            also_video=args.also_video,
            col_codigo=args.col_codigo,
            col_image_link=args.col_image_link,
            col_video_link=args.col_video_link,
            dic_links=links,
            sheet_name=args.sheet_name
        )

if __name__ == "__main__":

    linker = LinkInserter()
    linker.main()