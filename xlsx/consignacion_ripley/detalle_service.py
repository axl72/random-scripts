import system
from util import parsear_fecha
from pathlib import Path
import openpyxl
import pandas as pd
import traceback

detalle_sheet_columns_total = 23

def detalle_intek(workbook_path: Path, 
                            sheet_name: str = "DETALLE", 
                            start_row: int = 6,
                            styles: dict = None) -> tuple[list, list]:
    

    
    # 1. Diccionario limpio: Mapea directamente el ÍNDICE de columna (int) con el TÍTULO (str)
    # Q=17, R=18, S=19, T=20, U=21, V=22
    nuevas_columnas = {
        17: "MASTER",
        18: "Precio Master Total",
        19: "costo",
        20: "ASUME PROVEEDOR",
        21: "DESCTO. AUTORIZADO",
        22: "PAGO NETO (DEBEN PAGAR)",
        23: "POR REGULARIZAR"
    }

    try:
        # 2. Cargar y copiar la hoja
        wb = openpyxl.load_workbook(workbook_path)
        detalle_sheet = wb[sheet_name]

        # 3. Calcular e insertar columnas faltantes
        missing_columns = detalle_sheet_columns_total - detalle_sheet.max_column # Cuántas columnas faltan para llegar a 23
        if missing_columns > 0:
            # Insertamos todas las columnas vacías de una vez en el índice 17 (Columna Q)
            detalle_sheet.insert_cols(idx=17, amount=missing_columns)

        # ---------------------------------------------------------
        # 4. ESCRIBIR ENCABEZADOS (FILA 6) - Sin bucles anidados
        # ---------------------------------------------------------
        for col_idx, titulo in nuevas_columnas.items():
            celda_titulo = detalle_sheet.cell(row=6, column=col_idx)
            celda_titulo.value = titulo
            celda_titulo.fill = system.styles["fill_celeste"]
            celda_titulo.font = system.styles["font_negro_negrita"]
            celda_titulo.alignment = system.styles["alignment_centro_ajustado"]

        # ---------------------------------------------------------
        # 5. ESCRIBIR FÓRMULAS/DATOS (DESDE FILA 7 HASTA EL FINAL)
        # ---------------------------------------------------------
        ultima_fila = detalle_sheet.max_row
        for fila in range(7, ultima_fila + 1):
            try:
                # AQUÍ ES DONDE PONES TUS FÓRMULAS DE EXCEL.
                # Ejemplo: Te pongo fórmulas ficticias, TÚ debes cambiarlas por las reales.
                # Recuerda usar f-strings para que el número de fila ({fila}) sea dinámico.
                sku_ripley = str(detalle_sheet.cell(row=fila, column=5).value)  # el SKU está en la columna E (5)
                sku_ripley = sku_ripley.strip() if isinstance(sku_ripley, str) else None  # Limpiar espacios si es string
                string_fecha_venta = str(detalle_sheet.cell(row=fila, column=1).value)  # la fecha de venta está en la columna A (1)
                fecha_venta = parsear_fecha(string_fecha_venta, format="%Y-%m-%d %H:%M:%S")  # Verificamos que la fecha se pueda parsear correctamente
                precio_master = system.sku_intek_df.loc[sku_ripley, "PVP SUGERIDO"]
                costo = float(system.sku_intek_df.loc[sku_ripley, "VALOR VENTA"]) if sku_ripley in system.sku_intek_df.index else 0

                detalle_sheet.cell(row=fila, column=17).value = float(precio_master)  # Ejemplo para Col Q
                detalle_sheet.cell(row=fila, column=17).number_format = system.styles["formato_moneda"]  # Formato de número con 2 decimales y separador de miles

                detalle_sheet.cell(row=fila, column=18).value = f"=Q{fila}*G{fila}" # Ejemplo para Col R
                detalle_sheet.cell(row=fila, column=18).number_format = system.styles["formato_moneda"]  # Formato de número con 2 decimales y separador de miles

                unidad_venta = detalle_sheet.cell(row=fila, column=7).value  # Columna G (UNIDAD DE VENTA)
                detalle_sheet.cell(row=fila, column=19).value = costo*unidad_venta # Col S (COSTO TOTAL)

                rec_uni = 0
                descuento_autorizado = 0
                if fecha_venta >= system.periodo[0] and fecha_venta <= system.periodo[1]:

                    descuento_autorizado = float(system.ofertas_intek_df.loc[sku_ripley, "% DSCTO"] if sku_ripley in system.ofertas_intek_df.index else 0) # Esto debe ser None pero cuando es None arroja una excepcion
                    rec_uni = system.ofertas_intek_df.loc[sku_ripley, "RECO PROVEEDOR (SIN IGV)"] if sku_ripley in system.ofertas_intek_df.index else 0

                    
                detalle_sheet.cell(row=fila, column=21).value = descuento_autorizado         # Columna DESCTO. AUTORIZADO
                detalle_sheet.cell(row=fila, column=21).number_format = '0%'  # Formato de número con 2 decimales y separador de miles
                detalle_sheet.cell(row=fila, column=22).value = float(costo)*unidad_venta - float(rec_uni)*unidad_venta       # Columna PAGO NETO (DEBEN PAGAR)A
                detalle_sheet.cell(row=fila, column=22).number_format = system.styles["formato_moneda"]  # Formato de número con 2 decimales y separador de miles
                detalle_sheet.cell(row=fila, column=23).value = f"=V{fila} - L{fila}"  # Columna POR REGULARIZAR
                detalle_sheet.cell(row=fila, column=23).number_format = system.styles["formato_moneda"]  # Formato de número con 2 decimales y separador de miles
            
                detalle_sheet.cell(row=fila, column=20).value = f"=1 - V{fila}/S{fila}" # Columna Asume Proveedor
                detalle_sheet.cell(row=fila, column=20).number_format = '0%'  # Formato de número con 2 decimales y separador de miles
            except Exception as e:
                traceback.print_exc()


        # ---------------------------------------------------------
        # 6. EXTRAER DATOS PARA RETORNAR (Como lo necesitabas)
        # ---------------------------------------------------------
        columns = [cell.value for cell in detalle_sheet[6]] # Lee la fila 6
        data = []
        
        # iter_rows con values_only=True es mucho más rápido y devuelve solo los valores, no objetos celda
        for row in detalle_sheet.iter_rows(min_row=7, max_row=ultima_fila, values_only=True):
            # Filtramos filas completamente vacías
            if any(cell is not None for cell in row):
                # Limpiamos strings si es necesario
                clean_row = [cell.strip() if isinstance(cell, str) else cell for cell in row]
                data.append(clean_row)

        # 7. GUARDAR EL ARCHIVO
        wb.save(system.output_path)
        print("✅ Archivo procesado y guardado exitosamente como 'output.xlsx'")

        # 8. RETORNAR LOS DOS VALORES ESPERADOS (¡Aquí estaba tu error original!)
        return data, columns

    except Exception as e:
        traceback.print_exc()
        
        # Retornamos tupla vacía para evitar el error de "unpack" en el main
        return [], []