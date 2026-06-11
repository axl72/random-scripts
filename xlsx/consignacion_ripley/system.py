from pathlib import Path
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime
from util import load_database
from consignacion_parser import parsear_argumentos
import json

with open("config.json", "r", encoding="utf-8") as f:
    config = json.load(f)

args = parsear_argumentos()

# Estilos
styles = {
    "fill_celeste": PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid"),
    "fill_naranja": PatternFill(start_color="E26B0A", end_color="E26B0A", fill_type="solid"),
    "font_negro_negrita": Font(bold=True, color="000000", size=8),
    "font_blanco": Font(bold=False, color="FFFFFF", size=10),
    "alignment_centro_ajustado": Alignment(horizontal="center", vertical="center", wrap_text=True),
    "formato_moneda": '_("S/"* #,##0.00_);_("S/"* -#,##0.00_);_("S/"* "-"??_);_(@_)'
}

path_sku_intek = Path(config["path_sku_intek"])
path_ofertas_intek = Path(r"C:\Users\dberrospi\Desktop\INTEK\7. CLIENTES\RIPLEY\DATA\Ofertas\INTEK PERU - BAZAR JUGUETES - OFERTAS  - MAYO 2026.xlsx")
path_ofertas_intek = args.path_ofertas
liquidacion_path = Path(r"C:\Users\dberrospi\Desktop\INTEK\7. CLIENTES\RIPLEY\CONSIGNACION\CONSIGNACION 2026 6 JUNIO\INTEK PERU_01-31.xlsx")
liquidacion_path = args.path_liquidacion
output_path = Path("output.xlsx")
output_path = args.output

inicio_oferta = "2026-05-09"
fin_oferta = "2026-05-31"
inicio_oferta_dt = datetime.strptime(inicio_oferta, "%Y-%m-%d")
fin_oferta_dt = datetime.strptime(fin_oferta, "%Y-%m-%d")
periodo = (inicio_oferta_dt, fin_oferta_dt)
periodo = args.periodo

sku_intek_df = load_database(path_sku_intek, index="SKU", sheet_name="SKU")  # Carga y muestra el DataFrame para verificar que se lee correctamente
print(f"SKU Intek DataFrame cargado con {len(sku_intek_df)} filas.")
ofertas_intek_df = load_database(path_ofertas_intek, index="SKU RIPLEY")  # Carga y muestra el DataFrame para verificar que se lee correctamente
date_format = "%Y-%m-%d %H:%M:%S"
date_format = args.date_format