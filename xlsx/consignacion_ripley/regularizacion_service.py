from util import parsear_fecha
import openpyxl
import traceback
from pathlib import Path
import system

def regularizacion_intek(path_liquidacion):

    nuevas_columnas = {
        17: "COSTO (VALOR VENTA)",
        18: "COSTO TOTAL",
        19: "REC. UNI",
        20: "REC. TOTAL",
        21: "PAGO NETO (DEBEN PAGAR)",
        22: "DIFERENCIA"
    }

    wb = openpyxl.load_workbook(path_liquidacion)
    sheet = wb["DETALLE"]
    detalle_sheet = wb.copy_worksheet(sheet)
    detalle_sheet.title = "REGULARIZACION"

    missing_columns = 22 - sheet.max_column # La cantidad de columnas que deben haber en la hoja de Regularización es 22 (hasta la V)
    if missing_columns > 0:
        # Insertamos todas las columnas vacías de una vez en el índice 17 (Columna Q)
        detalle_sheet.insert_cols(idx=17, amount=missing_columns)

    # Escribir columnas nuevas (fila 6)
    for col_idx, titulo in nuevas_columnas.items():
        celda_titulo = detalle_sheet.cell(row=6, column=col_idx)
        celda_titulo.value = titulo
        celda_titulo.fill = system.styles["fill_naranja"]
        celda_titulo.font = system.styles["font_blanco"]
        celda_titulo.alignment = system.styles["alignment_centro_ajustado"]


    ultima_fila = detalle_sheet.max_row
    for fila in range(7, ultima_fila):
        if not fila:
            continue # Acá debería estar la suma de los totales de lo que estoy pidiendo de regularización
        try:
            sku_ripley = str(detalle_sheet.cell(row=fila, column=5).value)  # el SKU está en la columna E (5)
            sku_ripley = sku_ripley.strip() if isinstance(sku_ripley, str) else None  # Limpiar espacios si es string
            
            # precio_master = sku_intek_df.loc[sku_ripley, "PVP SUGERIDO"]
            costo = float(system.sku_intek_df.loc[sku_ripley, "VALOR VENTA"])
  
            string_fecha_venta = str(detalle_sheet.cell(row=fila, column=1).value)  # la fecha de venta está en la columna A (1)
            fecha_venta = parsear_fecha(string_fecha_venta, format=system.date_format)  # Verificamos que la fecha se pueda parsear correctamente


            detalle_sheet.cell(row=fila, column=17).value = costo  # Valor del costo (VALOR VENTA) para Col Q
            detalle_sheet.cell(row=fila, column=17).number_format = system.styles["formato_moneda"]  # Formato de número con 2 decimales y separador de miles
            detalle_sheet.cell(row=fila, column=18).value = f"=Q{fila}*G{fila}" # Ejemplo para Col R
            detalle_sheet.cell(row=fila, column=18).number_format = system.styles["formato_moneda"]  # Formato de número con 2 decimales y separador de miles

            if fecha_venta >= system.periodo[0] and fecha_venta <= system.periodo[1]:
                detalle_sheet.cell(row=fila, column=19).value = f"=IFERROR(VLOOKUP(E{fila}*1,Tabla[#All],8,0),0)" # Rec Unitario
                detalle_sheet.cell(row=fila, column=19).number_format = system.styles["formato_moneda"]  # Formato de número con 2 decimales y separador de miles
            detalle_sheet.cell(row=fila, column=20).value = f"=S{fila}*G{fila}"  # Rec Total
            detalle_sheet.cell(row=fila, column=20).number_format = system.styles["formato_moneda"]  # Formato de número con 2 decimales y separador de miles
            detalle_sheet.cell(row=fila, column=21).value = f"=R{fila}-T{fila}"  # Pago Neto (Deben Pagar)

            
            detalle_sheet.cell(row=fila, column=21).number_format = system.styles["formato_moneda"]  # Formato de número con 2 decimales y separador de miles
            detalle_sheet.cell(row=fila, column=22).value = f"=U{fila}-L{fila}"  # Diferencia entre lo que deben pagar y lo que se les pagará (POR REGULARIZAR)
            detalle_sheet.cell(row=fila, column=22).number_format = system.styles["formato_moneda"]  # Formato de número con 2 decimales y separador de miles
        except Exception as e:
            print(f"Error procesando fila {fila}, SKU: {sku_ripley}")
    wb.save("output.xlsx")
    return Path("output.xlsx")