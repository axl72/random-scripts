import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
import pandas as pd
from pathlib import Path

def procesar_ofertas_liquidacion(path_ofertas, path_liquidacion, path_salida = "output.xlsx"):
    """
    Lee una hoja de ofertas, la agrega como hoja nueva al libro de liquidación 
    (sin modificar el original) y guarda el resultado como un archivo nuevo.
    """
    nombre_tabla = "Tabla"
    nombre_hoja = "OFERTAS"

    # 1. Leer los datos de origen rápidamente con pandas
    print(f"Cargando Ofertas desde: {path_ofertas}")
    df = pd.read_excel(path_ofertas) # Que sea una lectura general para un archivo de una hoja, no necestimos el nombre
    df = df[["SKU RIPLEY",
            "DESCRIPCIÓN RIPLEY",
            "PRECIO MASTER",
            "PRECIO PROMO HOT SALES",
            "FECHA INICIO",
            "FECHA FIN",
            "STOCK UND",
            "RECO PROVEEDOR (SIN IGV)",
            "% DSCTO"
            ]]
    print(f"Hoja de Ofertas cargada con {len(df)} filas")
    # 2. Cargar el libro de liquidación ORIGINAL en memoria (el archivo en disco no se toca)
    print(f"Cargando Liquidacion Intek desde {path_liquidacion}")
    wb = openpyxl.load_workbook(path_liquidacion)
    print(f"Liquidación Intek cargada correctamente")
    
    # 3. Crear la hoja nueva en este libro cargado en memoria
    ws = wb.create_sheet(title=nombre_hoja)
    
    # 4. Insertar los datos del DataFrame en la nueva hoja (conservando encabezados)
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
        
    # 5. Convertir el rango de datos en una Tabla oficial de Excel
    max_fila = ws.max_row
    max_col = ws.max_column
    letra_col = get_column_letter(max_col)
    rango_tabla = f"A1:{letra_col}{max_fila}"
    
    nombre_tabla_limpio = nombre_tabla.replace(" ", "_")
    tabla = Table(displayName=nombre_tabla_limpio, ref=rango_tabla)
    
    # Estilo de tabla profesional con filas alternas
    estilo = TableStyleInfo(
        name="TableStyleMedium9", 
        showFirstColumn=False,
        showLastColumn=False, 
        showRowStripes=True,      
        showColumnStripes=False
    )
    tabla.tableStyleInfo = estilo
    ws.add_table(tabla)
    
    # 6. Guardar la fusión como un archivo TOTALMENTE NUEVO
    new_path =  path_salida
    wb.save(new_path)
    print("Ofertas agregadas correctamente al archivo de liquidación", new_path)
    return Path(new_path)  # Retornamos la ruta del nuevo archivo generado