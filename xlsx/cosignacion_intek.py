import argparse
import traceback
import openpyxl
from pathlib import Path
import pandas as pd
from datetime import datetime


# Primero se prepara la estructura (insertar columnas).
# Luego se pintan los títulos (fila 6).
# Luego se inyectan las fórmulas (fila 7 en adelante).
# Finalmente leo los datos.

detalle_sheet_columns_total = 23
detalle_sheet_title = "DETALLE_PROCESADO" # Cambié el nombre para evitar conflicto con la hoja original
path_sku_intek = Path("C:\\Users\\dberrospi\\Desktop\\.REPORTES\\Ripley.xlsx")
path_ofertas_intek = Path("C:\\Users\\dberrospi\\Desktop\\.REPORTES\\Ripley.xlsx")
inicio_oferta = "2026-05-09"
fin_oferta = "2026-05-31"
inicio_oferta_dt = datetime.strptime(inicio_oferta, "%Y-%m-%d")
fin_oferta_dt = datetime.strptime(fin_oferta, "%Y-%m-%d")
periodo = (inicio_oferta_dt, fin_oferta_dt)

def load_database(path: Path, index: str = None, sheet_name: str =  None) -> pd.DataFrame:
    if sheet_name:
        return  pd.read_excel(path, sheet_name=sheet_name, dtype=str)
    return pd.read_excel(path, dtype=str)


def read_consginacion_intek(workbook_path: Path, 
                            sheet_name: str = "DETALLE", 
                            start_row: int = 6) -> tuple[list, list]:
    
    sku_intek_df = load_database(path_sku_intek, "SKU", sheet_name="SKU")  # Carga y muestra el DataFrame para verificar que se lee correctamente
    ofertas_intek_df = load_database(path_ofertas_intek, index="SKU RIPLEY")  # Carga y muestra el DataFrame para verificar que se lee correctamente
    
    # 1. Diccionario limpio: Mapea directamente el ÍNDICE de columna (int) con el TÍTULO (str)
    # Q=17, R=18, S=19, T=20, U=21, V=22
    nuevas_columnas = {
        17: "MASTER",
        18: "Precio Master Total",
        19: "costo",
        20: "ASUME PROVEEDOR",
        21: "DESCTO. AUTORIZADO",
        22: "PAGO NETO (DEBEN PAGAR)",
        23: "POR REGULARIZAR"
    }

    try:
        # 2. Cargar y copiar la hoja
        wb = openpyxl.load_workbook(workbook_path)
        sheet = wb[sheet_name]
        detalle_sheet = wb.copy_worksheet(sheet)
        detalle_sheet.title = detalle_sheet_title

        # 3. Calcular e insertar columnas faltantes
        missing_columns = detalle_sheet_columns_total - sheet.max_column
        if missing_columns > 0:
            # Insertamos todas las columnas vacías de una vez en el índice 17 (Columna Q)
            detalle_sheet.insert_cols(idx=17, amount=missing_columns)

        # ---------------------------------------------------------
        # 4. ESCRIBIR ENCABEZADOS (FILA 6) - Sin bucles anidados
        # ---------------------------------------------------------
        for col_idx, titulo in nuevas_columnas.items():
            celda_titulo = detalle_sheet.cell(row=6, column=col_idx)
            celda_titulo.value = titulo
            
            # Copiar formato de la columna anterior (Columna 16 = 'P') para que se vea profesional
            # celda_referencia = detalle_sheet.cell(row=6, column=16)
            # if celda_referencia.font:
            #     celda_titulo.font = celda_referencia.font.copy()
            # if celda_referencia.fill:
            #     celda_titulo.fill = celda_referencia.fill.copy()

        # ---------------------------------------------------------
        # 5. ESCRIBIR FÓRMULAS/DATOS (DESDE FILA 7 HASTA EL FINAL)
        # ---------------------------------------------------------
        ultima_fila = detalle_sheet.max_row
          
        for fila in range(7, ultima_fila + 1):
            try:
                # AQUÍ ES DONDE PONES TUS FÓRMULAS DE EXCEL.
                # Ejemplo: Te pongo fórmulas ficticias, TÚ debes cambiarlas por las reales.
                # Recuerda usar f-strings para que el número de fila ({fila}) sea dinámico.
                sku_ripley = detalle_sheet.cell(row=fila, column=5).value  # el SKU está en la columna E (5)
                sku_ripley = sku_ripley.strip() if isinstance(sku_ripley, str) else None  # Limpiar espacios si es string
                string_fecha_venta = detalle_sheet.cell(row=fila, column=1).value  # la fecha de venta está en la columna A (1)
                day, mes, year = map(str.strip, string_fecha_venta.split("-"))
                mes = 5
                fecha_venta = datetime(int(year), int(mes), int(day))
                # mes = "5" if mes == "may" else mes  # Si el mes es "may", lo convertimos a "5"
                
                precio_master = sku_intek_df.loc[sku_ripley, "PVP SUGERIDO"]
                print(f"Fila {fila}: SKU RIPLEY='{sku_ripley}', Fecha Venta='{fecha_venta.strftime('%Y-%m-%d')}', Precio Master Encontrado='{precio_master}'")
                print(sku_intek_df.head())  # Mostrar las primeras filas del DataFrame para verificar que se cargó correctamente
                input("Presiona Enter para continuar...")  # Pausa para revisar cada fila
                costo = sku_intek_df.loc[sku_ripley, "VALOR VENTA"] if sku_ripley in sku_intek_df.index else None

                detalle_sheet.cell(row=fila, column=17).value = precio_master  # Ejemplo para Col Q
                detalle_sheet.cell(row=fila, column=18).value = f"=Q{fila}*G{fila}" # Ejemplo para Col R
                detalle_sheet.cell(row=fila, column=19).value = costo # Ejemplo para Col S
                if fecha_venta >= periodo[0] and fecha_venta <= periodo[1]:

                    descuento_autorizado = sku_intek_df.loc[sku_ripley, "% DSCTO"] if sku_ripley in sku_intek_df.index else None

                    detalle_sheet.cell(row=fila, column=21).value = descuento_autorizado         # Columna DESCTO. AUTORIZADO
                    detalle_sheet.cell(row=fila, column=22).value = costo - 1         # Columna PAGO NETO (DEBEN PAGAR)A
            
                detalle_sheet.cell(row=fila, column=20).value = f"=1 - V{fila}/S{fila}" # Columna Asume Proveedor
            except Exception as e:
                pass

        # ---------------------------------------------------------
        # 6. EXTRAER DATOS PARA RETORNAR (Como lo necesitabas)
        # ---------------------------------------------------------
        columns = [cell.value for cell in detalle_sheet[6]] # Lee la fila 6
        data = []
        
        # iter_rows con values_only=True es mucho más rápido y devuelve solo los valores, no objetos celda
        for row in detalle_sheet.iter_rows(min_row=7, max_row=ultima_fila, values_only=True):
            # Filtramos filas completamente vacías
            if any(cell is not None for cell in row):
                # Limpiamos strings si es necesario
                clean_row = [cell.strip() if isinstance(cell, str) else cell for cell in row]
                data.append(clean_row)

        # 7. GUARDAR EL ARCHIVO
        wb.save("output-consignacion.xlsx")
        print("✅ Archivo procesado y guardado exitosamente como 'output-consignacion.xlsx'")

        # 8. RETORNAR LOS DOS VALORES ESPERADOS (¡Aquí estaba tu error original!)
        return data, columns

    except Exception as e:
        traceback.print_exc()
        
        # Retornamos tupla vacía para evitar el error de "unpack" en el main
        return [], []

def main():
    path = Path(r"C:\Users\dberrospi\Desktop\INTEK\7. CLIENTES\RIPLEY\CONSIGNACION\CONSIGNACION 2026 5 MAYO\INTEK_01-30 - original.xlsx")
    
    # Ahora SÍ funcionará, porque la función siempre devuelve una tupla de 2 elementos
    data, columns = read_consginacion_intek(path)
    
    print(f"\nSe procesaron {len(data)} filas de datos.")
    print("Encabezados finales:", columns)

if __name__ == "__main__":
    main()