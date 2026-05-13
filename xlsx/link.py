def insertar_links_excel(
    archivo_entrada,
    archivo_salida,
    col_codigo="C",
    col_link="B",
    dic_links=None,
    link = "Link",
    hoja=0
):
    """
    Inserta hipervínculos comparando siempre como strings.
    """
    if dic_links is None:
        raise ValueError("Debes pasar un diccionario de links {codigo: url}")

    wb = load_workbook(archivo_entrada)
    ws = wb[wb.sheetnames[hoja]] if isinstance(hoja, int) else wb[hoja]

    for row in range(2, ws.max_row + 1):
        valor_celda = ws[f"{col_codigo}{row}"].value
        
        if valor_celda is None:
            continue

        # PARSEO CRÍTICO: Convertimos el valor de Excel a string y limpiamos espacios
        codigo_excel = str(valor_celda).strip()

        if codigo_excel in dic_links:
            url = dic_links[codigo_excel]

            cell = ws[f"{col_link}{row}"]
            cell.value = link  # Texto que se verá en la celda
            cell.hyperlink = url
            cell.font = Font(color="0000FF", underline="single")

    wb.save(archivo_salida)
    print(f"Archivo guardado exitosamente en: {archivo_salida}")

if __name__ == "__main__":
    from openpyxl import load_workbook
    from openpyxl.styles import Font

    # Ejemplo de uso
    dic_links = {
        "123": "http://example.com/123",
        "456": "http://example.com/456",
        "789": "http://example.com/789"
    }

    insertar_links_excel(
        archivo_entrada="input.xlsx",
        archivo_salida="output.xlsx",
        col_codigo="C",
        col_link="B",
        dic_links=dic_links,
        link="Ver Link",
        hoja=0
    )